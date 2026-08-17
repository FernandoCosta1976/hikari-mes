import json
import re
from collections import defaultdict
import openpyxl

BASE = "docs/prototype/source-data/business-validation/foundry/"
LINHA_C_FILE = BASE + "Quadro HH 2026 Linha C (2).xlsx"
PECAS_FILE = BASE + "PecasPorModelo.xlsx"

FAMILIES = ["CARC_ESQ", "CARC_DIR", "TAMPA_ESQ", "TAMPA_DIR", "CILINDRO", "CABECOTE"]
FAMILY_LABEL_PT = {
    "CARC_ESQ": "Carcaça Esquerda",
    "CARC_DIR": "Carcaça Direita",
    "TAMPA_ESQ": "Tampa Esquerda",
    "TAMPA_DIR": "Tampa Direita",
    "CILINDRO": "Cilindro",
    "CABECOTE": "Cabeçote",
}

# LINHA C OFC columns (1-indexed): A=1 Item, B=2 Modelo, C=3 Cor, D=4 Qtd, E=5 Lote,
# H=8 Data Real, I=9 Motor Hora Inicial, J=10 Motor Hora Final,
# AB=28 CarcDir, AC=29 CarcEsq, AD=30 TampaDir, AE=31 TampaEsq, AF=32 Cilindro, AG=33 Cabecote
LINHA_C_FAMILY_COLS = {
    "CARC_DIR": 28,
    "CARC_ESQ": 29,
    "TAMPA_DIR": 30,
    "TAMPA_ESQ": 31,
    "CILINDRO": 32,
    "CABECOTE": 33,
}

# FUNDIÇÃO columns: A=1 MOD.CORRENTE, C=3 Categoria, D=4 Descrição
# F=6 CarcEsq, G=7 CarcDir, H=8 TampaEsq, I=9 TampaDir, J=10 Cilindro, K=11 Cabecote(F)
FUNDICAO_FAMILY_COLS = {
    "CARC_ESQ": 6,
    "CARC_DIR": 7,
    "TAMPA_ESQ": 8,
    "TAMPA_DIR": 9,
    "CILINDRO": 10,
    "CABECOTE": 11,
}


def norm(v):
    if v is None:
        return ""
    return str(v).strip()


def norm_upper(v):
    return norm(v).upper()


def excel_time_str(v):
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        try:
            return v.strftime("%H:%M:%S")
        except Exception:
            return str(v)
    return str(v)


def excel_date_str(v):
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
    return str(v)


# ---------------------------------------------------------------------------
# 1. Load FUNDIÇÃO sheet -> resolve (prefix3, family) -> code | AMBIGUOUS | NOT_FOUND
# ---------------------------------------------------------------------------
wb_pecas = openpyxl.load_workbook(PECAS_FILE, data_only=True)
ws_fundicao = wb_pecas["FUNDIÇÃO"]

fundicao_rows = []
for row in range(2, ws_fundicao.max_row + 1):
    mod_corrente = ws_fundicao.cell(row=row, column=1).value
    if mod_corrente is None or norm(mod_corrente) == "":
        continue
    mod_corrente = norm(mod_corrente)
    prefix3 = mod_corrente.upper()[:3]
    category = norm(ws_fundicao.cell(row=row, column=3).value)
    description = norm(ws_fundicao.cell(row=row, column=4).value)
    families = {}
    for fam, col in FUNDICAO_FAMILY_COLS.items():
        raw = ws_fundicao.cell(row=row, column=col).value
        val = norm(raw)
        if val == "" or val == "-":
            families[fam] = None
        else:
            families[fam] = val.upper()
    fundicao_rows.append({
        "modCorrente": mod_corrente,
        "prefix3": prefix3,
        "category": category,
        "description": description,
        "families": families,
    })

prefix_groups = defaultdict(list)
for r in fundicao_rows:
    prefix_groups[r["prefix3"]].append(r)

# resolution[(prefix3, family)] = {"status": "RESOLVED"|"AMBIGUOUS"|"NOT_FOUND", "code": str|None, "candidates": [...]}
resolution = {}
for prefix3, rows in prefix_groups.items():
    for fam in FAMILIES:
        codes = set()
        for r in rows:
            c = r["families"].get(fam)
            if c:
                codes.add(c)
        if len(codes) == 0:
            resolution[(prefix3, fam)] = {"status": "NOT_FOUND", "code": None, "candidates": []}
        elif len(codes) == 1:
            resolution[(prefix3, fam)] = {"status": "RESOLVED", "code": next(iter(codes)), "candidates": list(codes)}
        else:
            resolution[(prefix3, fam)] = {"status": "AMBIGUOUS", "code": None, "candidates": sorted(codes)}


def resolve_component(prefix3, fam):
    if prefix3 not in prefix_groups:
        return {"status": "MODEL_MAPPING_NOT_FOUND", "code": None}
    res = resolution.get((prefix3, fam))
    if res is None or res["status"] == "NOT_FOUND":
        return {"status": "COMPONENT_MAPPING_NOT_FOUND", "code": None}
    if res["status"] == "AMBIGUOUS":
        return {"status": "MODEL_MAPPING_AMBIGUOUS", "code": None, "candidates": res["candidates"]}
    return {"status": "RESOLVED", "code": res["code"]}


# ---------------------------------------------------------------------------
# 2. Load "máquina titular e reserva" -> machine master
# ---------------------------------------------------------------------------
ws_machine = wb_pecas["máquina titular e reserva"]
machine_master = []  # list of {codigo, primary, primaryStandard, reserves:[...], reserveStandard}
for row in range(2, ws_machine.max_row + 1):
    codigo = ws_machine.cell(row=row, column=2).value
    if codigo is None or norm(codigo) == "":
        continue
    codigo = norm(codigo).upper().replace("\xa0", "")
    proposta = norm(ws_machine.cell(row=row, column=6).value)
    primary = norm(ws_machine.cell(row=row, column=7).value).upper()
    primary_standard = norm(ws_machine.cell(row=row, column=8).value).upper()
    reserve_raw = norm(ws_machine.cell(row=row, column=9).value).upper()
    reserve_standard = norm(ws_machine.cell(row=row, column=10).value).upper()
    reserves = []
    if reserve_raw and reserve_raw != "NÃO":
        for part in re.split(r"[/,]", reserve_raw):
            part = part.strip()
            if part and part != "NÃO":
                reserves.append(part)
    machine_master.append({
        "codigo": codigo,
        "proposta": proposta,
        "primary": primary if primary else None,
        "primaryStandard": primary_standard if primary_standard else None,
        "reserves": reserves,
        "reserveStandard": reserve_standard if reserve_standard else None,
    })


def dashless(code):
    return code.replace("-", "").upper()


def resolve_machine_alias(canonical_code):
    target = dashless(canonical_code)
    matches = [m for m in machine_master if m["codigo"].startswith(target)]
    if not matches:
        return {"status": "RESOURCE_MAPPING_NOT_FOUND", "entries": []}
    # verify consistency across matches (primary/reserves must agree)
    signatures = set()
    for m in matches:
        signatures.add((m["primary"], tuple(sorted(m["reserves"]))))
    if len(signatures) > 1:
        return {"status": "RESOURCE_MAPPING_AMBIGUOUS", "entries": matches}
    return {"status": "RESOLVED", "entries": matches}


# ---------------------------------------------------------------------------
# 3. Load LINHA C OFC and explode requirements
# ---------------------------------------------------------------------------
wb_linha = openpyxl.load_workbook(LINHA_C_FILE, data_only=True)
ws_linha = wb_linha["LINHA C OFC"]

total_source_rows = 0
rows_all_na = 0
rows_applicable = 0
requirements = []
exclusions = []

for row in range(5, ws_linha.max_row + 1):
    item = ws_linha.cell(row=row, column=1).value
    if item is None:
        continue
    total_source_rows += 1
    modelo = norm(ws_linha.cell(row=row, column=2).value)
    cor = norm(ws_linha.cell(row=row, column=3).value)
    qtde = ws_linha.cell(row=row, column=4).value
    lote = ws_linha.cell(row=row, column=5).value
    data_real = excel_date_str(ws_linha.cell(row=row, column=8).value)
    motor_hora_inicial = excel_time_str(ws_linha.cell(row=row, column=9).value)
    motor_hora_final = excel_time_str(ws_linha.cell(row=row, column=10).value)

    applicable_families = []
    for fam, col in LINHA_C_FAMILY_COLS.items():
        raw = ws_linha.cell(row=row, column=col).value
        val_norm = norm_upper(raw)
        if val_norm == "NA":
            continue
        applicable_families.append(fam)

    if not applicable_families:
        rows_all_na += 1
        continue
    rows_applicable += 1

    model_prefix = modelo.upper()[:3] if modelo else ""
    source_item = int(item) if isinstance(item, (int, float)) else norm(item)

    for fam in applicable_families:
        comp = resolve_component(model_prefix, fam)
        base_req = {
            "sourceRowIndex": row,
            "sourceItem": source_item,
            "sourceModel": modelo,
            "modelPrefix": model_prefix,
            "sourceColor": cor,
            "sourceQuantity": qtde,
            "sourceLot": lote,
            "sourceBusinessDate": data_real,
            "sourceDemandStart": motor_hora_inicial,
            "sourceDemandFinish": motor_hora_final,
            "family": fam,
        }
        if comp["status"] != "RESOLVED":
            exclusions.append({
                **base_req,
                "reason": comp["status"],
                "candidates": comp.get("candidates", []),
            })
            continue

        canonical_code = comp["code"]
        alias = resolve_machine_alias(canonical_code)
        if alias["status"] != "RESOLVED":
            exclusions.append({
                **base_req,
                "componentCode": canonical_code,
                "reason": alias["status"],
            })
            continue

        entry = alias["entries"][0]
        primary_raw = entry["primary"]
        reserves_raw = entry["reserves"]

        def normalize_resource(v):
            if v is None:
                return None
            v = v.strip().upper()
            m = re.match(r"^DC0?(\d)$", v)
            if m:
                return f"DC0{m.group(1)}"
            return v  # LP1..LP4 or other

        primary_resource = normalize_resource(primary_raw)
        reserve_resources = [normalize_resource(r) for r in reserves_raw]

        if primary_resource and primary_resource.startswith("DC"):
            productive_area = "FOUNDRY_DC"
        elif primary_resource and primary_resource.startswith("LP"):
            productive_area = "FOUNDRY_LP"
        elif entry["proposta"] and "LP" in entry["proposta"]:
            productive_area = "FOUNDRY_LP"
        else:
            productive_area = "UNKNOWN"

        if productive_area == "UNKNOWN":
            exclusions.append({
                **base_req,
                "componentCode": canonical_code,
                "reason": "PRODUCTIVE_AREA_UNRESOLVED",
            })
            continue

        requirements.append({
            **base_req,
            "componentCode": canonical_code,
            "primaryResource": primary_resource,
            "reserveResources": reserve_resources,
            "primaryStandardStatus": entry["primaryStandard"],
            "reserveStandardStatus": entry["reserveStandard"],
            "productiveArea": productive_area,
            "machineMasterCode": entry["codigo"],
        })

# ---------------------------------------------------------------------------
# 4. Aggregate + audit
# ---------------------------------------------------------------------------
dc_reqs = [r for r in requirements if r["productiveArea"] == "FOUNDRY_DC"]
lp_reqs = [r for r in requirements if r["productiveArea"] == "FOUNDRY_LP"]

exclusion_reason_counts = defaultdict(lambda: {"count": 0, "quantity": 0})
for e in exclusions:
    key = (e.get("modelPrefix", ""), e["family"], e["reason"])
    exclusion_reason_counts[key]["count"] += 1
    q = e.get("sourceQuantity") or 0
    exclusion_reason_counts[key]["quantity"] += q

total_exploded = len(requirements) + len(exclusions)

audit = {
    "A_linhas_consideradas_linha_c_ofc": total_source_rows,
    "B_linhas_ignoradas_all_na": rows_all_na,
    "linhas_aplicaveis": rows_applicable,
    "C_component_requirements_explodidos": total_exploded,
    "D_requirements_resolvidos": len(requirements),
    "E_ambiguidades": sum(1 for e in exclusions if e["reason"] in ("MODEL_MAPPING_AMBIGUOUS", "RESOURCE_MAPPING_AMBIGUOUS")),
    "F_unmatched": sum(1 for e in exclusions if e["reason"] in ("MODEL_MAPPING_NOT_FOUND", "COMPONENT_MAPPING_NOT_FOUND", "RESOURCE_MAPPING_NOT_FOUND", "PRODUCTIVE_AREA_UNRESOLVED")),
    "G_requirements_dc": len(dc_reqs),
    "H_requirements_lp": len(lp_reqs),
    "I_quantidade_total_dc": sum(r["sourceQuantity"] or 0 for r in dc_reqs),
    "J_quantidade_total_excluida": sum(e.get("sourceQuantity") or 0 for e in exclusions),
}

output = {
    "audit": audit,
    "requirements": requirements,
    "exclusions": exclusions,
    "fundicaoResolutionSample": {f"{k[0]}|{k[1]}": v for k, v in list(resolution.items())[:5]},
    "machineMasterCount": len(machine_master),
    "fundicaoRowCount": len(fundicao_rows),
    "prefixGroupCount": len(prefix_groups),
}

with open("scripts/data-pipeline/output.json", "w") as f:
    json.dump(output, f, indent=2, default=str, ensure_ascii=False)

print(json.dumps(audit, indent=2, ensure_ascii=False))
print("\nExclusion reasons breakdown:")
for k, v in sorted(exclusion_reason_counts.items()):
    print(f"  prefix={k[0]:5s} family={k[1]:10s} reason={k[2]:30s} count={v['count']:4d} qty={v['quantity']}")

# ---------------------------------------------------------------------------
# 5. Canonical component master (unique components actually used by resolved reqs)
# ---------------------------------------------------------------------------
components_by_code = {}
for r in requirements:
    code = r["componentCode"]
    if code not in components_by_code:
        components_by_code[code] = {
            "canonicalCode": code,
            "family": r["family"],
            "productiveArea": r["productiveArea"],
        }

component_resource_mappings = {}
component_aliases = []
for code, comp in components_by_code.items():
    alias = resolve_machine_alias(code)
    if alias["status"] == "RESOLVED":
        entry = alias["entries"][0]
        component_aliases.append({
            "canonicalComponentCode": code,
            "sourceCode": entry["codigo"],
            "source": "máquina titular e reserva",
            "matchMethod": "DASHLESS_PREFIX",
            "status": "RESOLVED",
        })

def normalize_resource2(v):
    if v is None:
        return None
    v = v.strip().upper()
    m = re.match(r"^DC0?(\d)$", v)
    if m:
        return f"DC0{m.group(1)}"
    return v

for m in machine_master:
    canonical_matches = [c for c in components_by_code if m["codigo"].startswith(dashless(c))]
    for code in canonical_matches:
        component_resource_mappings[code] = {
            "componentCode": code,
            "primaryResource": normalize_resource2(m["primary"]),
            "primaryStandardStatus": m["primaryStandard"],
            "reserveResources": [normalize_resource2(r) for r in m["reserves"]],
            "reserveStandardStatus": m["reserveStandard"],
            "machineMasterCode": m["codigo"],
        }

# ---------------------------------------------------------------------------
# 6. Full model x family resolution table (for audit / traceability)
# ---------------------------------------------------------------------------
model_component_mappings = []
for prefix3 in sorted(prefix_groups.keys()):
    for fam in FAMILIES:
        res = resolution.get((prefix3, fam))
        if res is None:
            continue
        model_component_mappings.append({
            "modelPrefix": prefix3,
            "family": fam,
            "status": res["status"],
            "resolvedCode": res["code"],
            "candidates": res["candidates"],
        })

# ---------------------------------------------------------------------------
# 7. 2026-07-09 operational subset (earliest business date -> demonstrative baseline)
# ---------------------------------------------------------------------------
OPERATIONAL_DATE = "2026-07-09"
operational_reqs = [r for r in requirements if r["sourceBusinessDate"] == OPERATIONAL_DATE and r["productiveArea"] == "FOUNDRY_DC"]
operational_lp_reqs = [r for r in requirements if r["sourceBusinessDate"] == OPERATIONAL_DATE and r["productiveArea"] == "FOUNDRY_LP"]

output2 = {
    "components": list(components_by_code.values()),
    "componentResourceMappings": list(component_resource_mappings.values()),
    "componentAliases": component_aliases,
    "modelComponentMappings": model_component_mappings,
    "operationalDate": OPERATIONAL_DATE,
    "operationalRequirementsDc": operational_reqs,
    "operationalRequirementsLp": operational_lp_reqs,
    "exclusionSummary": [
        {"modelPrefix": k[0], "family": k[1], "reason": k[2], "count": v["count"], "quantity": v["quantity"]}
        for k, v in sorted(exclusion_reason_counts.items())
    ],
}

with open("scripts/data-pipeline/output2.json", "w") as f:
    json.dump(output2, f, indent=2, default=str, ensure_ascii=False)

print("\n\n=== components ===", len(components_by_code))
print("=== component_resource_mappings ===", len(component_resource_mappings))
print("=== component_aliases ===", len(component_aliases))
print("=== operational (2026-07-09) DC reqs ===", len(operational_reqs), "LP reqs:", len(operational_lp_reqs))
print("operational DC source lots:", sorted(set(r['sourceLot'] for r in operational_reqs)))
print("operational total qty (DC, unique lot):")
seen_lots=set()
total=0
for r in operational_reqs:
    if r['sourceLot'] not in seen_lots:
        seen_lots.add(r['sourceLot'])
